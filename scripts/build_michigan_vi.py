"""Michigan Vote Intent margins, read from the dropped-in workbook rather than SQL.

    python scripts/build_michigan_vi.py
    python scripts/build_michigan_vi.py --dry-run
    python scripts/build_michigan_vi.py --workbook "path/to/other.xlsx"

Source: data/Michigan Vote Intent District Margins.xlsx, sheet "VoteIntent2026" — one row
per district, labelled "SD: n" (senate) or "LD: n" (house), plus a statewide row that is
skipped. The same workbook also carries HighMid2026 / All26Targets sheets, but those two
variants are built from SQL by build_model_margins.py; only Vote Intent lives here.

Vote Intent is modelled separately and is NOT reproducible from universe counts, which is
why it cannot come from the RSLC_MI_R2 exchange table like All and H+M do.

"netframework" is already Republican minus Democrat as a fraction — the GOP-minus-Dem
convention used everywhere else — and just needs scaling to percentage points. app.js
negates the rslc family on display to reach the Dem-positive convention.

Writes only model_rslc_vi. It deliberately leaves model_rslc_all / model_rslc_hm alone, so
it can run before or after build_model_margins.py (whose MI entry lists model_rslc_vi under
`preserve` for the same reason).

Run this AFTER generate_chamber_jsons.py, which rebuilds chamber files from the election
workbook and drops anything written here.
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from build_model_margins import DATA_DIR, chamber_path, seg_key

WORKBOOK = DATA_DIR / "Michigan Vote Intent District Margins.xlsx"
SHEET = "VoteIntent2026"

FAMILY = "rslc"
FAMILY_LABEL = "RSLC"
VARIANT = "vi"
VARIANT_LABEL = "VI"
VIEW = f"model_{FAMILY}_{VARIANT}"

NAME_COL = 1        # "tabname" — "SD: 12"; column 0 is a numeric tabkey
NET_COL = 4         # "netframework", Republican minus Democrat as a fraction
FIRST_UNIVERSE = 8  # "univ_RepBase" .. "univ_DemBase"

# The workbook abbreviates the universe ladder; the exchange table spells it out. Map one
# to the other so a VI breakdown carries the same segment keys and order as the All / H+M
# breakdowns built from SQL — otherwise the same universe renders under two different keys
# and picks up a different colour from the family palette.
UNIVERSE_COLUMNS = [
    ("univ_RepBase", "Republican Base"),
    ("univ_RepTargs", "Republican Targets"),
    ("univ_TrumpOver", "2024 Trump Overperform"),
    ("univ_MsgTarg", "Message Targets"),
    ("univ_StubMiddle", "Stubborn Middle"),
    ("univ_AvailDem", "Available Dems"),
    ("univ_DemTargs", "Democrat Targets"),
    ("univ_DemBase", "Democrat Base"),
]

DISTRICT_RE = re.compile(r"^\s*(SD|LD)\s*:\s*(\d+)\s*$", re.I)
CHAMBER_OF = {"SD": "senate", "LD": "house"}


def check_header(header, workbook):
    """The column order is positional, so a reshaped workbook must fail loudly rather
    than silently mapping Republican Base counts onto Democrat Base."""
    got = [str(h).strip() if h is not None else "" for h in header]
    expected = [c for c, _label in UNIVERSE_COLUMNS]
    actual = got[FIRST_UNIVERSE:FIRST_UNIVERSE + len(expected)]
    problems = []
    if got[NET_COL] != "netframework":
        problems.append(f"column {NET_COL} is {got[NET_COL]!r}, expected 'netframework'")
    if actual != expected:
        problems.append(f"universe columns are {actual}, expected {expected}")
    if problems:
        sys.exit(f"{workbook.name} / {SHEET} has an unexpected layout:\n  "
                 + "\n  ".join(problems))


def read_sheet(ws, workbook):
    """chamber -> {district_id -> (segments, margin)}"""
    rows = ws.iter_rows(values_only=True)
    check_header(next(rows), workbook)

    out = {"house": {}, "senate": {}}
    for row in rows:
        if not row or row[NAME_COL] is None:
            continue
        m = DISTRICT_RE.match(str(row[NAME_COL]))
        if not m:
            continue          # statewide total row
        chamber = CHAMBER_OF[m.group(1).upper()]
        district_id = f"{int(m.group(2)):03d}"

        net = row[NET_COL]
        if net is None:
            continue
        segments = [
            {"key": seg_key(label), "label": label,
             "value": round(100.0 * float(row[FIRST_UNIVERSE + i] or 0), 1)}
            for i, (_col, label) in enumerate(UNIVERSE_COLUMNS)
        ]
        out[chamber][district_id] = (segments, round(100.0 * float(net), 1))
    return out


def patch(chamber, table, dry_run=False):
    path = chamber_path("MI", chamber)
    recs = json.loads(path.read_text(encoding="utf-8"))
    hits = 0
    for rec in recs:
        key = str(rec.get("district_id"))
        if key not in table:
            # A district that resolved on a previous run but not on this one must not
            # keep the old number.
            (rec.get("view_margins") or {}).pop(VIEW, None)
            (rec.get("models") or {}).pop(VIEW, None)
            continue
        segments, margin = table[key]
        affinity = {"segments": segments, "margin": margin}
        for s in segments:
            affinity[s["key"]] = s["value"]
        rec.setdefault("view_margins", {})[VIEW] = margin
        rec.setdefault("models", {})[VIEW] = {
            "label": f"{FAMILY_LABEL} ({VARIANT_LABEL})",
            "family": FAMILY_LABEL,
            "variant": VARIANT_LABEL,
            "affinity": affinity,
        }
        hits += 1

    if not dry_run:
        path.write_text(json.dumps(recs, indent=1, ensure_ascii=False), encoding="utf-8")

    sample = ", ".join(f"{d}:{m:+.1f}" for d, (_s, m) in sorted(table.items())[:3])
    flag = "" if hits == len(recs) else "  <-- INCOMPLETE"
    print(f"  {chamber:<7} {hits:>3}/{len(recs):<4} districts   {sample}{flag}"
          + ("   (dry run)" if dry_run else ""))
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", help=f"Override the source workbook (default: {WORKBOOK.name})")
    ap.add_argument("--dry-run", action="store_true", help="Compute but do not write files")
    args = ap.parse_args()

    workbook = Path(args.workbook) if args.workbook else WORKBOOK
    if not workbook.exists():
        sys.exit(f"Missing {workbook}")

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Workbook is missing sheet {SHEET!r}; found {wb.sheetnames}")

    results = read_sheet(wb[SHEET], workbook)
    print(f"Michigan  RSLC (VI)  [{workbook.name} / {SHEET}]")
    for chamber in ("house", "senate"):
        patch(chamber, results[chamber], dry_run=args.dry_run)


if __name__ == "__main__":
    main()
