"""Kansas model margins, read from the dropped-in workbook rather than SQL.

    python scripts/build_kansas_margins.py
    python scripts/build_kansas_margins.py --dry-run

Source: data/Kansas Model.xlsx, two tabs —
    All26Targets   -> the "All" variant   (Targetable 2026 Electorate)
    HighMid2026    -> the "H+M" variant   (High/Mid Prop 2026)

Each tab has one row per district labelled "SD: n" (senate) or "LD: n" (house), plus a
statewide total row that is skipped. "Net Framework" is already Kobach minus Democrat as a
fraction, which matches the GOP-minus-Dem convention used everywhere else — it just needs
scaling to percentage points. The nine "Universe: ..." columns become the affinity
breakdown and sum to 1.0.

Kansas has no state-specific model in SQL, so build_model_margins.EXTERNAL_MODELS lists it
to keep the national fallback from claiming the state.
"""

import argparse
import json
import re
import sys

import openpyxl

from build_model_margins import DATA_DIR, chamber_path, seg_key

WORKBOOK = DATA_DIR / "Kansas Model.xlsx"

SHEETS = {"all": "All26Targets", "hm": "HighMid2026"}
VARIANT_LABEL = {"all": "All", "hm": "H+M"}

# The workbook's framework columns are named for Kobach, but the model is published as RAGA.
FAMILY = "raga"
FAMILY_LABEL = "RAGA"

NET_COL = 2          # "Net Framework"
FIRST_UNIVERSE = 6   # "Universe: Kobach Base"

DISTRICT_RE = re.compile(r"^\s*(SD|LD)\s*:\s*(\d+)\s*$", re.I)
CHAMBER_OF = {"SD": "senate", "LD": "house"}


def read_sheet(ws):
    """(chamber -> {district_id -> (segments, margin)})"""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    labels = []
    for h in header[FIRST_UNIVERSE:]:
        if h is None:
            break
        labels.append(re.sub(r"^\s*Universe:\s*", "", str(h)).strip())

    out = {"house": {}, "senate": {}}
    for row in rows:
        if not row or row[0] is None:
            break
        m = DISTRICT_RE.match(str(row[0]))
        if not m:
            continue          # statewide total row
        chamber = CHAMBER_OF[m.group(1).upper()]
        district_id = f"{int(m.group(2)):03d}"

        net = row[NET_COL]
        if net is None:
            continue
        segments = []
        for i, label in enumerate(labels):
            v = row[FIRST_UNIVERSE + i]
            segments.append({"key": seg_key(label), "label": label,
                             "value": round(100.0 * float(v or 0), 1)})
        out[chamber][district_id] = (segments, round(100.0 * float(net), 1))
    return out


def patch(chamber, results, dry_run=False):
    path = chamber_path("KS", chamber)
    recs = json.loads(path.read_text(encoding="utf-8"))
    produced = {f"model_{FAMILY}_{v}" for v in results}
    hits = 0
    for rec in recs:
        key = str(rec.get("district_id"))
        vm = rec.setdefault("view_margins", {})
        models = rec.setdefault("models", {})

        # Kansas previously took the national fallback; a dedicated model supersedes it.
        for stale in [k for k in vm if k.startswith("model_drnatl_")
                      or (k.startswith(f"model_{FAMILY}_") and k not in produced)]:
            vm.pop(stale, None)
        for stale in [k for k in models if k.startswith("model_drnatl_")
                      or (k.startswith(f"model_{FAMILY}_") and k not in produced)]:
            models.pop(stale, None)

        matched = False
        for variant, table in results.items():
            if key not in table:
                continue
            segments, margin = table[key]
            affinity = {"segments": segments, "margin": margin}
            for s in segments:
                affinity[s["key"]] = s["value"]
            view = f"model_{FAMILY}_{variant}"
            vm[view] = margin
            models[view] = {
                "label": f"{FAMILY_LABEL} ({VARIANT_LABEL[variant]})",
                "family": FAMILY_LABEL,
                "variant": VARIANT_LABEL[variant],
                "affinity": affinity,
            }
            matched = True
        if matched:
            hits += 1
        if not models:
            rec.pop("models", None)

    if not dry_run:
        path.write_text(json.dumps(recs, indent=1, ensure_ascii=False), encoding="utf-8")
    flag = "" if hits == len(recs) else "  <-- INCOMPLETE"
    print(f"  {chamber:<7} {hits}/{len(recs)}{flag}" + ("   (dry run)" if dry_run else ""))
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not WORKBOOK.exists():
        sys.exit(f"Missing {WORKBOOK}")

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    missing = [s for s in SHEETS.values() if s not in wb.sheetnames]
    if missing:
        sys.exit(f"Workbook is missing sheet(s): {missing}; found {wb.sheetnames}")

    per_variant = {v: read_sheet(wb[s]) for v, s in SHEETS.items()}
    print(f"Kansas  [{WORKBOOK.name}]  variants={','.join(SHEETS)}")
    for chamber in ("house", "senate"):
        results = {v: per_variant[v][chamber] for v in SHEETS}
        patch(chamber, results, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
